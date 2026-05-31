from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel, Field
import openpyxl
from pymongo import MongoClient
import os
import io
import uvicorn
import datetime
from dotenv import load_dotenv

# Load environment variables from .env
load_dotenv()

# MongoDB Configuration
MONGODB_URI = os.getenv("MONGODB_URI")
if not MONGODB_URI:
    raise ValueError("MONGODB_URI environment variable is missing in .env file!")

# Initialize MongoDB Connection
try:
    mongo_client = MongoClient(MONGODB_URI, serverSelectionTimeoutMS=5000)
    # Trigger connection check
    mongo_client.server_info()
    db = mongo_client["credit_db"]
    clients_col = db["clients"]
    print("✅ تم الاتصال بنجاح بقاعدة بيانات MongoDB Atlas!")
except Exception as e:
    print(f"❌ خطأ أثناء الاتصال بقاعدة بيانات MongoDB Atlas: {e}")
    raise SystemExit("فشل الاتصال بقاعدة بيانات MongoDB! يرجى التحقق من اتصال الإنترنت والإعدادات.")

# Create FastAPI app
app = FastAPI(title="Credit Risk Analyzer API", version="2.0.0")

# Allow CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "نظام_تحليل_العملاء_الائتماني.xlsx")

# Pydantic Client model for request validation
class ClientData(BaseModel):
    name: str = Field(..., description="Client Name")
    id_num: str = Field(..., description="National ID")
    mobile: str = Field(..., description="Mobile Number")
    age: int = Field(..., description="Age")
    employer: str = Field(..., description="Employer")
    emp_type: str = Field(..., description="Employment Type (حكومي, شبه حكومي, خاص كبير, خاص صغير, متقاعد)")
    basic_sal: float = Field(..., description="Basic Salary")
    gross_sal: float = Field(..., description="Gross Salary")
    net_sal: float = Field(..., description="Net Salary")
    svc_months: int = Field(..., description="Service Months")
    simah: int = Field(..., description="SIMAH Score")
    inquiries: int = Field(..., description="90 Days Inquiries")
    default_status: str = Field(..., description="Active Default (نعم/لا)")
    blacklist: str = Field(..., description="Service Suspended (نعم/لا)")
    sal_attach: str = Field(..., description="Salary Attached (نعم/لا)")
    
    # Debts
    loans_count: int = Field(0)
    loans_total: float = Field(0.0)
    real_estate_count: int = Field(0)
    real_estate_total: float = Field(0.0)
    cards_count: int = Field(0)
    cards_total: float = Field(0.0)
    finance_cos_count: int = Field(0)
    finance_cos_total: float = Field(0.0)
    monthly_installment: float = Field(0.0)
    
    # Executions
    exec_requests_count: int = Field(0)
    exec_requests_total: float = Field(0.0)
    ind_exec_count: int = Field(0)
    ind_exec_total: float = Field(0.0)
    fin_exec_count: int = Field(0)
    fin_exec_total: float = Field(0.0)
    bank_exec_count: int = Field(0)
    bank_exec_total: float = Field(0.0)
    
    # Fee %
    fees_percent: float = Field(0.10)
    
    # Workflow Stage
    workflow_stage: str = Field("جديد", description="Workflow Stage")

    # Workflow Approvals
    analyst_name: str = Field("", description="Analyst Name")
    analyst_date: str = Field("", description="Analyst Date")
    analyst_recommendation: str = Field("", description="Analyst Recommendation")
    
    ops_name: str = Field("", description="Operations Name")
    ops_date: str = Field("", description="Operations Date")
    ops_decision: str = Field("", description="Operations Decision")
    
    finance_name: str = Field("", description="Finance Name")
    finance_date: str = Field("", description="Finance Date")
    finance_decision: str = Field("", description="Finance Decision")
    
    general_notes: str = Field("", description="General Notes")
    completion_date: str = Field("", description="Completion Date")


def calculate_client_metrics(c: dict) -> dict:
    """Calculates all credit risk and feasibility metrics matching Excel sheet formulas"""
    # Ensure net_sal is strictly calculated as: Gross Salary - 9% of Basic Salary
    basic_sal = float(c.get("basic_sal") or 0.0)
    gross_sal = float(c.get("gross_sal") or 0.0)
    c["net_sal"] = gross_sal - (basic_sal * 0.09)
    
    # 1. DTI
    dti = c["monthly_installment"] / c["net_sal"] if c["net_sal"] > 0 else 0.0
    
    # 2. SIMAH Points (Max 25)
    simah = c["simah"]
    if simah >= 750: simah_pts = 25
    elif simah >= 700: simah_pts = 20
    elif simah >= 650: simah_pts = 15
    elif simah >= 600: simah_pts = 10
    elif simah >= 550: simah_pts = 5
    else: simah_pts = 0
    
    # 3. DTI Points (Max 25)
    if dti <= 0.3: dti_pts = 25
    elif dti <= 0.4: dti_pts = 20
    elif dti <= 0.5: dti_pts = 15
    elif dti <= 0.6: dti_pts = 10
    elif dti <= 0.7: dti_pts = 5
    else: dti_pts = 0
    
    # 4. Defaults Points (Max 20)
    if c["default_status"] == "نعم" or c["blacklist"] == "نعم" or c["sal_attach"] == "نعم":
        default_pts = 0
    else:
        default_pts = 20
        
    # 5. Executions Points (Max 20)
    tot_exec_count = c["exec_requests_count"] + c["ind_exec_count"] + c["fin_exec_count"] + c["bank_exec_count"]
    tot_exec_val = c["exec_requests_total"] + c["ind_exec_total"] + c["fin_exec_total"] + c["bank_exec_total"]
    
    if tot_exec_count == 0: exec_pts = 20
    elif tot_exec_count > 8 or tot_exec_val > 100000: exec_pts = 0
    elif tot_exec_count > 5 or tot_exec_val > 50000: exec_pts = 5
    elif tot_exec_count > 3 or tot_exec_val > 25000: exec_pts = 10
    else: exec_pts = 15
    
    # 6. Employer Points (Max 15)
    emp_type = c["emp_type"]
    if emp_type == "حكومي": emp_pts = 15
    elif emp_type == "شبه حكومي": emp_pts = 10
    elif emp_type == "خاص كبير": emp_pts = 8
    elif emp_type == "متقاعد": emp_pts = 5
    else: emp_pts = 5
    
    # Total Points (Max 105)
    total_pts = simah_pts + dti_pts + default_pts + exec_pts + emp_pts
    score_pct = (total_pts / 105.0) * 100
    
    # Classification
    if total_pts >= 85: classification = "A"
    elif total_pts >= 70: classification = "B"
    elif total_pts >= 55: classification = "C"
    else: classification = "D"
    
    # Direct rejection triggers & reasons
    reasons = []
    if c["default_status"] == "نعم": reasons.append("تعثر قائم")
    if c["blacklist"] == "نعم": reasons.append("إيقاف خدمات")
    if c["sal_attach"] == "نعم": reasons.append("حجز راتب")
    if tot_exec_count > 8: reasons.append("تجاوز الحد الأعلى للتنفيذات")
    if tot_exec_val > 100000: reasons.append("تجاوز الحد الأقصى لمبالغ التنفيذات")
    if c["fin_exec_count"] > 3: reasons.append("تنفيذات تمويلية مرتفعة (شركات تمويل)")
    if simah < 550: reasons.append("درجة سمة منخفضة (أقل من 550)")
    if dti > 0.7: reasons.append("استقطاع مرتفع (يتجاوز 70%)")
    
    # Final Decision
    if reasons:
        decision = "مرفوض"
        brief_reasons = "سياسة الرفض المباشر: " + "، ".join(reasons)
    else:
        if classification == "A":
            decision = "مؤهل"
            brief_reasons = "استيفاء شروط الجدارة الائتمانية الممتازة (A)"
        elif classification == "B":
            decision = "مؤهل بتحفظ"
            brief_reasons = "استيفاء الشروط مع درجة مخاطر متوسطة (B)"
        elif classification == "C":
            decision = "يحتاج استثناء"
            brief_reasons = "مخاطر عالية تحتاج استثناء إداري خاص (C)"
        else:
            decision = "مرفوض"
            brief_reasons = "النقاط الائتمانية الكلية منخفضة جداً (D)"
            
    # Risk Level (based on Executions)
    if tot_exec_count == 0:
        risk_level = "لا توجد تنفيذات"
    elif tot_exec_count > 8 or tot_exec_val > 100000 or c["fin_exec_count"] > 3:
        risk_level = "مرفوض مباشر"
    elif tot_exec_count > 5 or tot_exec_val > 50000:
        risk_level = "مخاطر عالية"
    elif tot_exec_count > 3 or tot_exec_val > 25000:
        risk_level = "مخاطر متوسطة"
    else:
        risk_level = "مخاطر منخفضة"
        
    # Feasibility
    if classification == "A": funding_factor = 20
    elif classification == "B": funding_factor = 15
    elif classification == "C": funding_factor = 10
    else: funding_factor = 0
    
    expected_funding = c["net_sal"] * funding_factor
    total_debts = c["loans_total"] + c["real_estate_total"] + c["cards_total"] + c["finance_cos_total"]
    total_executions = tot_exec_val
    required_payment = total_debts + total_executions
    
    gross_surplus = expected_funding - required_payment
    company_fees = expected_funding * c["fees_percent"]
    net_surplus = gross_surplus - company_fees
    
    if net_surplus > 50000:
        feasibility_decision = "فرصة ممتازة"
    elif net_surplus > 20000:
        feasibility_decision = "فرصة جيدة"
    elif net_surplus > 0:
        feasibility_decision = "فرصة ضعيفة"
    else:
        feasibility_decision = "غير مجدي اقتصادياً"
        
    return {
        "file_id": c.get("file_id", "ملف-جديد"),
        "date": c.get("date", datetime.date.today().strftime("%Y-%m-%d")),
        "name": c["name"],
        "id_num": c["id_num"],
        "mobile": c["mobile"],
        "age": c["age"],
        "employer": c["employer"],
        "emp_type": c["emp_type"],
        "basic_sal": c["basic_sal"],
        "gross_sal": c["gross_sal"],
        "net_sal": c["net_sal"],
        "svc_months": c["svc_months"],
        "simah": c["simah"],
        "inquiries": c["inquiries"],
        "default_status": c["default_status"],
        "blacklist": c["blacklist"],
        "sal_attach": c["sal_attach"],
        
        "loans_count": c["loans_count"],
        "loans_total": c["loans_total"],
        "real_estate_count": c["real_estate_count"],
        "real_estate_total": c["real_estate_total"],
        "cards_count": c["cards_count"],
        "cards_total": c["cards_total"],
        "finance_cos_count": c["finance_cos_count"],
        "finance_cos_total": c["finance_cos_total"],
        "total_debts": total_debts,
        "monthly_installment": c["monthly_installment"],
        "dti_pct": round(dti * 100, 1),
        
        "exec_requests_count": c["exec_requests_count"],
        "exec_requests_total": c["exec_requests_total"],
        "ind_exec_count": c["ind_exec_count"],
        "ind_exec_total": c["ind_exec_total"],
        "fin_exec_count": c["fin_exec_count"],
        "fin_exec_total": c["fin_exec_total"],
        "bank_exec_count": c["bank_exec_count"],
        "bank_exec_total": c["bank_exec_total"],
        "total_exec_count": tot_exec_count,
        "total_exec_val": tot_exec_val,
        "risk_level": risk_level,
        
        "simah_pts": simah_pts,
        "dti_pts": dti_pts,
        "default_pts": default_pts,
        "exec_pts": exec_pts,
        "emp_pts": emp_pts,
        "total_pts": total_pts,
        "score_pct": round(score_pct, 1),
        "classification": classification,
        "decision": decision,
        "brief_reasons": brief_reasons,
        
        "expected_funding": expected_funding,
        "required_payment": required_payment,
        "fees_percent": c["fees_percent"],
        "company_fees": company_fees,
        "gross_surplus": gross_surplus,
        "net_surplus": net_surplus,
        "feasibility_decision": feasibility_decision,
        
        "workflow_stage": c.get("workflow_stage", "جديد"),
        "analyst_name": c.get("analyst_name", ""),
        "analyst_date": c.get("analyst_date", ""),
        "analyst_recommendation": c.get("analyst_recommendation", ""),
        "ops_name": c.get("ops_name", ""),
        "ops_date": c.get("ops_date", ""),
        "ops_decision": c.get("ops_decision", ""),
        "finance_name": c.get("finance_name", ""),
        "finance_date": c.get("finance_date", ""),
        "finance_decision": c.get("finance_decision", ""),
        "general_notes": c.get("general_notes", ""),
        "completion_date": c.get("completion_date", ""),
        "row_idx": c.get("row_idx", -1)
    }


# ─────────────────────────────────────────────
# MongoDB CRUD Helper Functions
# ─────────────────────────────────────────────

def db_get_clients() -> list:
    """Retrieves all clients from MongoDB, sorted by row_idx"""
    try:
        clients = list(clients_col.find({}).sort("row_idx", 1))
        for c in clients:
            c.pop("_id", None)
        return clients
    except Exception as e:
        print(f"خطأ أثناء جلب البيانات من MongoDB: {e}")
        return []


def db_save_client(client_data: dict) -> dict:
    """Saves a new client or updates an existing one in MongoDB"""
    existing = clients_col.find_one({"id_num": client_data["id_num"]})
    
    if existing:
        row_idx = existing["row_idx"]
        file_id = existing.get("file_id") or f"ملف-{row_idx - 3:03d}"
        client_data["row_idx"] = row_idx
        client_data["file_id"] = file_id
        client_data["date"] = existing.get("date", datetime.date.today().strftime("%Y-%m-%d"))
        calculated = calculate_client_metrics(client_data)
        clients_col.replace_one({"_id": existing["_id"]}, calculated)
        return calculated
    else:
        all_clients = list(clients_col.find({}, {"row_idx": 1}))
        if all_clients:
            max_idx = max(c["row_idx"] for c in all_clients)
            row_idx = max(max_idx + 1, 4)
        else:
            row_idx = 4
            
        file_id = f"ملف-{row_idx - 3:03d}"
        client_data["row_idx"] = row_idx
        client_data["file_id"] = file_id
        client_data["date"] = datetime.date.today().strftime("%Y-%m-%d")
        
        calculated = calculate_client_metrics(client_data)
        clients_col.insert_one(calculated)
        calculated.pop("_id", None)
        return calculated


def db_delete_client(row_idx: int) -> bool:
    """Deletes a client document from MongoDB by row_idx"""
    res = clients_col.delete_one({"row_idx": row_idx})
    return res.deleted_count > 0


# ─────────────────────────────────────────────
# Excel Import / Export Helpers
# ─────────────────────────────────────────────

def read_clients_from_excel_bytes(excel_bytes: bytes) -> list:
    """Reads all clients from an Excel file (given as bytes) using openpyxl"""
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    sh1 = wb['1. بيانات العميل']
    sh2 = wb['2. الالتزامات']
    sh3 = wb['3. التنفيذات']
    sh5 = wb['5. الجدوى التمويلية']
    sh6 = wb['6. الاعتمادات']
    
    clients_list = []
    for r in range(4, 54):
        name = sh1[f"C{r}"].value
        if not name or name == 0 or name == "0":
            continue
            
        file_id = sh1[f"A{r}"].value or f"ملف-{r-3:03d}"
        date_val = sh1[f"B{r}"].value
        if hasattr(date_val, "strftime"):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = str(date_val) if date_val else datetime.date.today().strftime("%Y-%m-%d")
            
        c = {
            "row_idx": r,
            "file_id": str(file_id),
            "date": date_str,
            "name": str(name),
            "id_num": str(sh1[f"D{r}"].value or ""),
            "mobile": str(sh1[f"E{r}"].value or ""),
            "age": int(sh1[f"F{r}"].value or 0),
            "employer": str(sh1[f"G{r}"].value or ""),
            "emp_type": str(sh1[f"H{r}"].value or "حكومي"),
            "basic_sal": float(sh1[f"I{r}"].value or 0.0),
            "gross_sal": float(sh1[f"J{r}"].value or 0.0),
            "net_sal": float(sh1[f"K{r}"].value or 0.0),
            "svc_months": int(sh1[f"L{r}"].value or 0),
            "simah": int(sh1[f"M{r}"].value or 0),
            "inquiries": int(sh1[f"N{r}"].value or 0),
            "default_status": str(sh1[f"O{r}"].value or "لا"),
            "blacklist": str(sh1[f"P{r}"].value or "لا"),
            "sal_attach": str(sh1[f"Q{r}"].value or "لا"),
            
            "loans_count": int(sh2[f"C{r}"].value or 0),
            "loans_total": float(sh2[f"D{r}"].value or 0.0),
            "real_estate_count": int(sh2[f"E{r}"].value or 0),
            "real_estate_total": float(sh2[f"F{r}"].value or 0.0),
            "cards_count": int(sh2[f"G{r}"].value or 0),
            "cards_total": float(sh2[f"H{r}"].value or 0.0),
            "finance_cos_count": int(sh2[f"I{r}"].value or 0),
            "finance_cos_total": float(sh2[f"J{r}"].value or 0.0),
            "monthly_installment": float(sh2[f"L{r}"].value or 0.0),
            
            "exec_requests_count": int(sh3[f"C{r}"].value or 0),
            "exec_requests_total": float(sh3[f"D{r}"].value or 0.0),
            "ind_exec_count": int(sh3[f"E{r}"].value or 0),
            "ind_exec_total": float(sh3[f"F{r}"].value or 0.0),
            "fin_exec_count": int(sh3[f"G{r}"].value or 0),
            "fin_exec_total": float(sh3[f"H{r}"].value or 0.0),
            "bank_exec_count": int(sh3[f"I{r}"].value or 0),
            "bank_exec_total": float(sh3[f"J{r}"].value or 0.0),
            
            "fees_percent": float(sh5[f"J{r}"].value or 0.10),
            "workflow_stage": str(sh1[f"R{r}"].value or "جديد"),
            
            "analyst_name": str(sh6[f"F{r}"].value or ""),
            "analyst_date": str(sh6[f"G{r}"].value or ""),
            "analyst_recommendation": str(sh6[f"H{r}"].value or ""),
            "ops_name": str(sh6[f"I{r}"].value or ""),
            "ops_date": str(sh6[f"J{r}"].value or ""),
            "ops_decision": str(sh6[f"K{r}"].value or ""),
            "finance_name": str(sh6[f"L{r}"].value or ""),
            "finance_date": str(sh6[f"M{r}"].value or ""),
            "finance_decision": str(sh6[f"N{r}"].value or ""),
            "general_notes": str(sh6[f"O{r}"].value or ""),
            "completion_date": str(sh6[f"P{r}"].value or "")
        }
        clients_list.append(calculate_client_metrics(c))
        
    return clients_list


def read_clients_from_excel_legacy() -> list:
    """Reads all clients from the local Excel file (used for initial migration)"""
    if not os.path.exists(EXCEL_PATH):
        return []
    with open(EXCEL_PATH, "rb") as f:
        return read_clients_from_excel_bytes(f.read())


def generate_excel_from_mongodb() -> bytes:
    """Generates a fresh Excel file from MongoDB data and returns it as bytes"""
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"ملف Excel الأساسي غير موجود: {EXCEL_PATH}")
    
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    sh1 = wb['1. بيانات العميل']
    sh2 = wb['2. الالتزامات']
    sh3 = wb['3. التنفيذات']
    sh5 = wb['5. الجدوى التمويلية']
    sh6 = wb['6. الاعتمادات']
    
    # Clear rows 4 to 53 first
    for r in range(4, 54):
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
            sh1[f"{col}{r}"] = None
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'L']:
            sh2[f"{col}{r}"] = None
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            sh3[f"{col}{r}"] = None
        sh5[f"J{r}"] = None
        for col in ['F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']:
            sh6[f"{col}{r}"] = None
            
    # Get live clients from MongoDB
    clients = db_get_clients()
    
    # Write clients to Excel
    for c in clients:
        r = c["row_idx"]
        if r < 4 or r > 53:
            continue
            
        sh1[f"A{r}"] = f'=IF(C{r}<>"","ملف-"&TEXT(ROW()-3,"000"),"")'
        sh1[f"B{r}"] = f'=IF(C{r}<>"",TODAY(),"")'
        sh1[f"C{r}"] = c["name"]
        sh1[f"D{r}"] = c["id_num"]
        sh1[f"E{r}"] = c["mobile"]
        sh1[f"F{r}"] = c["age"]
        sh1[f"G{r}"] = c["employer"]
        sh1[f"H{r}"] = c["emp_type"]
        sh1[f"I{r}"] = c["basic_sal"]
        sh1[f"J{r}"] = c["gross_sal"]
        sh1[f"K{r}"] = f"=J{r}-(I{r}*0.09)"
        sh1[f"L{r}"] = c["svc_months"]
        sh1[f"M{r}"] = c["simah"]
        sh1[f"N{r}"] = c["inquiries"]
        sh1[f"O{r}"] = c["default_status"]
        sh1[f"P{r}"] = c["blacklist"]
        sh1[f"Q{r}"] = c["sal_attach"]
        sh1[f"R{r}"] = c.get("workflow_stage", "جديد")
        
        sh2[f"A{r}"] = f"='1. بيانات العميل'!A{r}"
        sh2[f"B{r}"] = f"='1. بيانات العميل'!C{r}"
        sh2[f"C{r}"] = c["loans_count"]
        sh2[f"D{r}"] = c["loans_total"]
        sh2[f"E{r}"] = c["real_estate_count"]
        sh2[f"F{r}"] = c["real_estate_total"]
        sh2[f"G{r}"] = c["cards_count"]
        sh2[f"H{r}"] = c["cards_total"]
        sh2[f"I{r}"] = c["finance_cos_count"]
        sh2[f"J{r}"] = c["finance_cos_total"]
        sh2[f"K{r}"] = f"=IFERROR(D{r}+F{r}+H{r}+J{r},0)"
        sh2[f"L{r}"] = c["monthly_installment"]
        sh2[f"M{r}"] = f"='1. بيانات العميل'!K{r}"
        sh2[f"N{r}"] = f"=IFERROR(IF(M{r}>0,L{r}/M{r},0),0)"
        
        sh3[f"A{r}"] = f"='1. بيانات العميل'!A{r}"
        sh3[f"B{r}"] = f"='1. بيانات العميل'!C{r}"
        sh3[f"C{r}"] = c["exec_requests_count"]
        sh3[f"D{r}"] = c["exec_requests_total"]
        sh3[f"E{r}"] = c["ind_exec_count"]
        sh3[f"F{r}"] = c["ind_exec_total"]
        sh3[f"G{r}"] = c["fin_exec_count"]
        sh3[f"H{r}"] = c["fin_exec_total"]
        sh3[f"I{r}"] = c["bank_exec_count"]
        sh3[f"J{r}"] = c["bank_exec_total"]
        sh3[f"K{r}"] = f"=IFERROR(C{r}+E{r}+G{r}+I{r},0)"
        sh3[f"L{r}"] = f"=IFERROR(D{r}+F{r}+H{r}+J{r},0)"
        sh3[f"M{r}"] = f'=IF(K{r}=0,"لا توجد تنفيذات",IF(OR(K{r}>8,L{r}>100000,G{r}>3),"مرفوض مباشر",IF(OR(K{r}>5,L{r}>50000),"مخاطر عالية",IF(OR(K{r}>3,L{r}>25000),"مخاطر متوسطة","مخاطر منخفضة"))))'
        
        sh5[f"A{r}"] = f"='1. بيانات العميل'!A{r}"
        sh5[f"B{r}"] = f"='1. بيانات العميل'!C{r}"
        sh5[f"C{r}"] = f"='4. محرك المخاطر'!R{r}"
        sh5[f"D{r}"] = f"='1. بيانات العميل'!K{r}"
        sh5[f"E{r}"] = f'=IF(C{r}="A",20,IF(C{r}="B",15,IF(C{r}="C",10,0)))'
        sh5[f"F{r}"] = f"=IFERROR(D{r}*E{r},0)"
        sh5[f"G{r}"] = f"='2. الالتزامات'!K{r}"
        sh5[f"H{r}"] = f"='3. التنفيذات'!L{r}"
        sh5[f"I{r}"] = f"=IFERROR(G{r}+H{r},0)"
        sh5[f"J{r}"] = c["fees_percent"]
        sh5[f"K{r}"] = f"=IFERROR(F{r}*J{r},0)"
        sh5[f"L{r}"] = f"=IFERROR(F{r}-I{r},0)"
        sh5[f"M{r}"] = f"=IFERROR(L{r}-K{r},0)"
        sh5[f"N{r}"] = f'=IF(B{r}="","",IF(M{r}>50000,"فرصة ممتازة",IF(M{r}>20000,"فرصة جيدة",IF(M{r}>0,"فرصة ضعيفة","غير مجدي اقتصادياً"))))'
        
        sh6[f"A{r}"] = f"='1. بيانات العميل'!A{r}"
        sh6[f"B{r}"] = f"='1. بيانات العميل'!C{r}"
        sh6[f"C{r}"] = f"='4. محرك المخاطر'!R{r}"
        sh6[f"D{r}"] = f"='4. محرك المخاطر'!S{r}"
        sh6[f"E{r}"] = f"='1. بيانات العميل'!R{r}"
        sh6[f"F{r}"] = c.get("analyst_name", "")
        sh6[f"G{r}"] = c.get("analyst_date", "")
        sh6[f"H{r}"] = c.get("analyst_recommendation", "")
        sh6[f"I{r}"] = c.get("ops_name", "")
        sh6[f"J{r}"] = c.get("ops_date", "")
        sh6[f"K{r}"] = c.get("ops_decision", "")
        sh6[f"L{r}"] = c.get("finance_name", "")
        sh6[f"M{r}"] = c.get("finance_date", "")
        sh6[f"N{r}"] = c.get("finance_decision", "")
        sh6[f"O{r}"] = c.get("general_notes", "")
        sh6[f"P{r}"] = c.get("completion_date", "")
    
    # Save to bytes buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────
# Data Migration from Excel to MongoDB on Startup
# ─────────────────────────────────────────────

def migrate_excel_to_mongodb():
    """Migrates existing clients from Excel to MongoDB if MongoDB is empty"""
    try:
        if clients_col.count_documents({}) == 0:
            print("📦 قاعدة بيانات MongoDB فارغة. جاري ترحيل البيانات من ملف Excel...")
            excel_clients = read_clients_from_excel_legacy()
            if excel_clients:
                clients_col.insert_many(excel_clients)
                print(f"✅ تم ترحيل {len(excel_clients)} ملف عميل بنجاح من Excel إلى MongoDB!")
            else:
                print("ℹ️  ملف Excel لا يحتوي على أي ملفات عملاء صالحة. سيتم بدء قاعدة البيانات فارغة.")
        else:
            count = clients_col.count_documents({})
            print(f"✅ تم العثور على {count} ملف عميل في MongoDB. تم تخطي الترحيل.")
            
            # Update/recalculate all existing clients in MongoDB to ensure the new net_sal formula is applied
            print("🔄 جاري تحديث جميع السجلات في MongoDB لتطبيق معادلة الراتب الصافي الجديدة...")
            all_clients = list(clients_col.find({}))
            for c in all_clients:
                updated = calculate_client_metrics(c)
                clients_col.replace_one({"_id": c["_id"]}, updated)
            print("✅ تم تحديث جميع الملفات وتطبيق المعادلة بنجاح!")
    except Exception as e:
        print(f"⚠️  خطأ أثناء ترحيل البيانات أو تحديثها: {e}")


# Call migration on server startup
migrate_excel_to_mongodb()


# ─────────────────────────────────────────────
# API Endpoints
# ─────────────────────────────────────────────

@app.get("/api/clients")
def get_clients():
    """List all clients with computed metrics from MongoDB"""
    try:
        return db_get_clients()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"حدث خطأ أثناء قراءة البيانات: {e}")


@app.get("/api/stats")
def get_stats():
    """Return administrative summary dashboard indicators from MongoDB"""
    try:
        clients = db_get_clients()
        
        num_files = len(clients)
        num_qualified = sum(1 for c in clients if c.get("decision") == "مؤهل")
        num_reserved = sum(1 for c in clients if c.get("decision") == "مؤهل بتحفظ")
        num_exceptions = sum(1 for c in clients if c.get("decision") == "يحتاج استثناء")
        num_rejected = sum(1 for c in clients if c.get("decision") == "مرفوض")
        
        total_debts = sum(c.get("total_debts", 0.0) for c in clients)
        total_executions = sum(c.get("total_exec_val", 0.0) for c in clients)
        total_surpluses = sum(max(0.0, c.get("net_surplus", 0.0)) for c in clients)
        total_fees = sum(c.get("company_fees", 0.0) for c in clients)
        total_funding = sum(c.get("expected_funding", 0.0) for c in clients)
        avg_risk_score = sum(c.get("total_pts", 0) for c in clients) / num_files if num_files > 0 else 0.0
        
        return {
            "num_files": num_files,
            "num_qualified": num_qualified + num_reserved,
            "num_fully_qualified": num_qualified,
            "num_reserved_qualified": num_reserved,
            "num_exceptions": num_exceptions,
            "num_rejected": num_rejected,
            "total_debts": total_debts,
            "total_executions": total_executions,
            "total_surpluses": total_surpluses,
            "total_fees": total_fees,
            "total_funding": total_funding,
            "avg_risk_score": round(avg_risk_score, 1)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"حدث خطأ أثناء احتساب المؤشرات: {e}")


@app.post("/api/clients")
def save_client(client: ClientData):
    """Save a new client or update an existing one in MongoDB"""
    try:
        c_dict = client.model_dump()
        saved = db_save_client(c_dict)
        return {
            "success": True,
            "message": "تم حفظ ملف العميل بنجاح في قاعدة بيانات MongoDB!",
            "client": saved
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"حدث خطأ أثناء الحفظ: {e}")


@app.delete("/api/clients/{row_idx}")
def delete_client(row_idx: int):
    """Delete a client from MongoDB by row_idx"""
    try:
        deleted = db_delete_client(row_idx)
        if not deleted:
            raise HTTPException(status_code=404, detail="لم يتم العثور على الملف الائتماني المطلوب حذفه!")
        return {"success": True, "message": "تم حذف الملف الائتماني بنجاح من قاعدة البيانات!"}
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"حدث خطأ أثناء الحذف: {e}")


@app.get("/api/download")
def download_excel():
    """Generate a fresh Excel file from MongoDB data and serve it for download"""
    try:
        excel_bytes = generate_excel_from_mongodb()
        from fastapi.responses import Response
        from urllib.parse import quote
        safe_filename = quote("نظام_تحليل_العملاء_الائتماني.xlsx", safe="")
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{safe_filename}"
            }
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"حدث خطأ أثناء تصدير ملف Excel: {e}")


@app.post("/api/upload")
async def upload_excel(file: UploadFile = File(...)):
    """Upload an Excel file, clear MongoDB, and import all records from the file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="يرجى تحميل ملف Excel صالح (.xlsx, .xls)")
    try:
        content = await file.read()
        
        # Validate the file first
        try:
            openpyxl.load_workbook(io.BytesIO(content))
        except Exception as ve:
            raise HTTPException(status_code=400, detail=f"ملف Excel غير صالح أو تالف: {ve}")
        
        # Read clients from the uploaded file
        uploaded_clients = read_clients_from_excel_bytes(content)
        
        # Clear MongoDB collection and import fresh data
        clients_col.delete_many({})
        if uploaded_clients:
            clients_col.insert_many(uploaded_clients)
            
        # Also save the file locally as the new template
        with open(EXCEL_PATH, "wb") as f:
            f.write(content)
            
        return {
            "success": True,
            "message": f"تم استيراد {len(uploaded_clients)} ملف عميل بنجاح من ملف Excel إلى قاعدة بيانات MongoDB!",
            "imported_count": len(uploaded_clients)
        }
    except HTTPException as he:
        raise he
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"حدث خطأ أثناء تحميل الملف: {e}")


# ─────────────────────────────────────────────
# Static File Serving
# ─────────────────────────────────────────────

@app.get("/")
def serve_home():
    return FileResponse("index.html")

app.mount("/", StaticFiles(directory="."), name="static")

if __name__ == "__main__":
    print("🚀 خادم الائتمان المالي يعمل الآن على الرابط: http://127.0.0.1:8000")
    uvicorn.run("server:app", host="127.0.0.1", port=8000, reload=True)
